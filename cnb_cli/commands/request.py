# cnb_cli/commands/request.py

import questionary
from pathlib import Path
from openpyxl import Workbook, load_workbook
from rich.console import Console
from dotenv import load_dotenv
import os

console = Console()
load_dotenv()  # Load .env variables

def get_env_value(key: str, default: str = "") -> str:
    """Get value from .env or fallback to default"""
    return os.getenv(key, default)

def get_excel_path(environment: str, job_name: str) -> Path:
    """Generate Excel file path based on environment and job"""
    safe_job = job_name.replace(" ", "_")
    return Path(f"{environment}_{safe_job}.xlsx")

def save_request_to_excel(file_path: Path, data: dict):
    """Save request data to Excel"""
    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Add header row
        ws.append(["Type", "Environment", "Job", "Path", "Username", "Password"])

    ws.append([
        data.get("type"),
        data.get("environment"),
        data.get("job"),
        data.get("path"),
        data.get("username"),
        data.get("password"),
    ])

    wb.save(file_path)
    console.print(f"[green]✅ Request saved to {file_path}[/green]")

def request_menu():
    console.print("\n🎫 [bold cyan]Request Management System[/bold cyan]\n")

    req_type = questionary.select(
        "What type of request would you like to make?",
        choices=[
            "Add Config Map",
            "Delete Config Map",
            "Update Config Map",
        ]
    ).ask()

    environment = questionary.select(
        "Select target environment",
        choices=["DEV", "UAT", "PROD"]
    ).ask()

    job = questionary.text("Enter job name").ask()

    path_default = str(Path.cwd())  # Default to current working directory
    path = questionary.text("Enter path (optional)", default=path_default).ask()

    # Auto-load from .env
    username_default = get_env_value("USERNAME", "")
    password_default = get_env_value("PASSWORD", "")

    username = questionary.text("Enter username", default=username_default).ask()
    password = questionary.password("Enter password", default=password_default).ask()

    request_data = {
        "type": req_type,
        "environment": environment,
        "job": job,
        "path": path,
        "username": username,
        "password": password,
    }

    # Show summary and confirm
    console.print("\n[cyan]Request Summary:[/cyan]")
    for k, v in request_data.items():
        console.print(f"{k}: {v}")

    confirm = questionary.confirm("Save this request to Excel?", default=True).ask()
    if confirm:
        excel_file = get_excel_path(environment, job)
        save_request_to_excel(excel_file, request_data)
